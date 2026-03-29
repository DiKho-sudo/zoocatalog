"""
Тестовый скрипт для просмотра структуры Excel файла
"""
import openpyxl
import sys

if len(sys.argv) < 2:
    print("Использование: python test_excel.py путь_к_файлу.xlsx")
    sys.exit(1)

file_path = sys.argv[1]
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

print(f"Файл: {file_path}")
print(f"Всего строк: {sheet.max_row}")
print(f"Всего колонок: {sheet.max_column}")
print("\n" + "="*80)

# Показываем первые 15 строк
for row_num in range(1, min(16, sheet.max_row + 1)):
    print(f"\nСтрока {row_num}:")
    row_values = [cell.value for cell in sheet[row_num]]
    for idx, value in enumerate(row_values, 1):
        if value:
            print(f"  Колонка {idx}: {value}")

