"""
Команда для импорта товаров из 1С Розница
Использование: python manage.py import_from_1c путь/к/файлу.xlsx
"""

from django.core.management.base import BaseCommand, CommandError
from django.core.files import File
from django.utils.text import slugify
from products.models import Product, Category, Brand, AnimalType
import openpyxl
import os
from decimal import Decimal
from pathlib import Path


class Command(BaseCommand):
    help = 'Импорт товаров из файла Excel, выгруженного из 1С Розница'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Путь к Excel файлу из 1С')
        parser.add_argument(
            '--images-dir',
            type=str,
            default=None,
            help='Папка с изображениями товаров (по артикулу)'
        )
        parser.add_argument(
            '--update',
            action='store_true',
            help='Обновлять существующие товары'
        )
        parser.add_argument(
            '--header-row',
            type=int,
            default=None,
            help='Номер строки с заголовками (если не указано - автопоиск)'
        )

    def handle(self, *args, **options):
        file_path = options['file_path']
        images_dir = options['images_dir']
        update_existing = options['update']
        header_row_param = options['header_row']

        if not os.path.exists(file_path):
            raise CommandError(f'Файл не найден: {file_path}')

        self.stdout.write(self.style.SUCCESS(f'Начинаем импорт из {file_path}'))

        # Загружаем Excel файл
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
        except Exception as e:
            raise CommandError(f'Ошибка при чтении файла: {e}')

        # Статистика
        created_count = 0
        updated_count = 0
        skipped_count = 0
        error_count = 0

        # Определяем строку с заголовками
        if header_row_param:
            # Используем явно указанную строку
            header_row = header_row_param
            self.stdout.write(f'Используем заголовки из строки: {header_row} (указано вручную)')
        else:
            # Автоматический поиск заголовков
            header_row = 1
            
            # Ищем строку где есть "Номенклатура" или "Наименование" или "Артикул"
            for row_num in range(1, min(15, sheet.max_row + 1)):
                row_values = [cell.value for cell in sheet[row_num]]
                for cell_value in row_values:
                    if cell_value and isinstance(cell_value, str):
                        val_lower = cell_value.lower()
                        if any(word in val_lower for word in ['номенклатура', 'наименование', 'артикул', 'товар']):
                            header_row = row_num
                            break
                if header_row > 1:
                    break
            
            self.stdout.write(f'Заголовки найдены автоматически в строке: {header_row}')
        
        # Получаем заголовки
        headers = [cell.value for cell in sheet[header_row]]
        self.stdout.write(f'Найдены колонки: {headers}')

        # Определяем индексы колонок
        col_map = self._map_columns(headers)

        # Создаем дефолтные объекты если их нет
        default_animal_type = self._get_or_create_default_animal_type()
        default_category = self._get_or_create_default_category()

        # Обрабатываем строки (начиная со следующей после заголовков)
        for row_num, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            try:
                product_data = self._extract_product_data(row, col_map, headers)
                
                if not product_data.get('name'):
                    self.stdout.write(self.style.WARNING(
                        f'Строка {row_num}: Пропущена (нет названия)'
                    ))
                    skipped_count += 1
                    continue

                # Получаем или создаем категорию
                category = self._get_or_create_category(
                    product_data.get('category'),
                    default_category
                )

                # Получаем или создаем бренд
                brand = self._get_or_create_brand(product_data.get('brand'))

                # Проверяем существование товара по артикулу
                article = product_data.get('article', '')
                existing_product = None
                
                if article:
                    existing_product = Product.objects.filter(
                        slug=slugify(article)
                    ).first()

                if existing_product and update_existing:
                    # Обновляем существующий товар
                    self._update_product(existing_product, product_data, category, brand, default_animal_type)
                    updated_count += 1
                    self.stdout.write(self.style.SUCCESS(
                        f'Строка {row_num}: Обновлен "{product_data["name"]}"'
                    ))
                elif not existing_product:
                    # Создаем новый товар
                    product = self._create_product(
                        product_data, 
                        category, 
                        brand, 
                        default_animal_type,
                        images_dir
                    )
                    created_count += 1
                    self.stdout.write(self.style.SUCCESS(
                        f'Строка {row_num}: Создан "{product_data["name"]}"'
                    ))
                else:
                    skipped_count += 1
                    self.stdout.write(self.style.WARNING(
                        f'Строка {row_num}: Пропущен "{product_data["name"]}" (уже существует)'
                    ))

            except Exception as e:
                error_count += 1
                self.stdout.write(self.style.ERROR(
                    f'Строка {row_num}: Ошибка - {str(e)}'
                ))

        # Выводим итоги
        self.stdout.write(self.style.SUCCESS('\n' + '='*50))
        self.stdout.write(self.style.SUCCESS('ИТОГИ ИМПОРТА:'))
        self.stdout.write(self.style.SUCCESS(f'Создано новых товаров: {created_count}'))
        self.stdout.write(self.style.SUCCESS(f'Обновлено товаров: {updated_count}'))
        self.stdout.write(self.style.WARNING(f'Пропущено товаров: {skipped_count}'))
        if error_count > 0:
            self.stdout.write(self.style.ERROR(f'Ошибок: {error_count}'))
        self.stdout.write(self.style.SUCCESS('='*50))

    def _map_columns(self, headers):
        """Определяет соответствие колонок"""
        col_map = {}
        
        # Возможные варианты названий колонок из 1С
        mappings = {
            'name': ['наименование', 'название', 'товар', 'номенклатура'],
            'article': ['артикул', 'код', 'article'],
            'price': ['цена', 'цена продажи', 'розничная цена', 'price'],
            'category': ['группа', 'категория', 'группа номенклатуры', 'category', 'вид номенклатуры'],
            'brand': ['производитель', 'бренд', 'торговая марка', 'brand', 'марка'],
            'stock': ['остаток', 'количество', 'stock', 'остатки'],
            'weight': ['вес', 'масса', 'weight'],
            'description': ['описание', 'комментарий', 'description', 'наименование полное'],
            'unit': ['единица измерения', 'единица', 'ед.изм'],
        }

        for idx, header in enumerate(headers):
            if not header:
                continue
            header_lower = str(header).lower().strip()
            for key, variants in mappings.items():
                if any(variant in header_lower for variant in variants):
                    col_map[key] = idx
                    break

        return col_map

    def _extract_product_data(self, row, col_map, headers):
        """Извлекает данные товара из строки"""
        data = {}
        
        for key, idx in col_map.items():
            if idx < len(row):
                value = row[idx]
                if value is not None:
                    data[key] = str(value).strip() if not isinstance(value, (int, float, Decimal)) else value
        
        # Извлекаем единицу измерения из заголовка или названия
        data['unit'] = 'шт'  # По умолчанию
        
        # Пытаемся извлечь из заголовка "Номенклатура, Характеристика, Ед. изм."
        if col_map.get('name') is not None and col_map['name'] < len(headers):
            header = str(headers[col_map['name']]).lower()
            if 'ед. изм' in header or 'ед.изм' in header:
                # Единица измерения может быть в самом названии товара
                name = data.get('name', '')
                # Ищем , шт или , кг в конце
                if ', шт' in name or ' шт' in name:
                    data['unit'] = 'шт'
                elif ', кг' in name or ' кг' in name or ', кг,' in name:
                    data['unit'] = 'кг'
                elif ', г' in name or ' г' in name:
                    data['unit'] = 'г'
                elif ', л' in name or ' л' in name:
                    data['unit'] = 'л'
                elif ', мл' in name or ' мл' in name:
                    data['unit'] = 'мл'

        return data

    def _get_or_create_default_animal_type(self):
        """Создает дефолтный тип животного"""
        animal_type, created = AnimalType.objects.get_or_create(
            slug='vse',
            defaults={
                'name': 'Все животные'
            }
        )
        return animal_type

    def _get_or_create_default_category(self):
        """Создает дефолтную категорию"""
        category, created = Category.objects.get_or_create(
            slug='imported',
            defaults={
                'name': 'Импортированные товары',
                'description': 'Товары, импортированные из 1С'
            }
        )
        return category

    def _get_or_create_category(self, category_name, default_category):
        """Получает или создает категорию"""
        if not category_name:
            return default_category

        slug = slugify(category_name)
        category, created = Category.objects.get_or_create(
            slug=slug,
            defaults={'name': category_name}
        )
        return category

    def _get_or_create_brand(self, brand_name):
        """Получает или создает бренд"""
        if not brand_name:
            brand_name = 'Неизвестный производитель'

        brand, created = Brand.objects.get_or_create(
            name=brand_name,
            defaults={'country': 'Беларусь'}
        )
        return brand

    def _create_product(self, data, category, brand, animal_type, images_dir):
        """Создает новый товар"""
        # Используем артикул если есть, иначе название
        article = data.get('article', '').strip()
        name = data['name']
        slug = slugify(article if article else name)
        
        # Если slug уже существует, добавляем счетчик
        base_slug = slug
        counter = 1
        while Product.objects.filter(slug=slug).exists():
            slug = f"{base_slug}-{counter}"
            counter += 1
        
        # Определяем статус наличия
        stock_status = 'in_stock'
        stock = data.get('stock')
        if stock is not None:
            try:
                stock_value = float(stock)
                if stock_value <= 0:
                    stock_status = 'out_of_stock'
            except (ValueError, TypeError):
                pass

        # Если нет цены, ставим 0 (можно обновить позже)
        price = self._parse_decimal(data.get('price', 0))
        if price == 0:
            self.stdout.write(self.style.WARNING(
                f'  ⚠ Цена не указана, установлена 0 BYN'
            ))

        product = Product.objects.create(
            name=name,
            slug=slug,
            description=data.get('description', '')[:500] if data.get('description') else '',
            composition='',
            weight=self._parse_decimal(data.get('weight', 1.0)),
            unit=data.get('unit', 'шт'),
            price=price,
            category=category,
            brand=brand,
            animal_type=animal_type,
            stock_status=stock_status,
            age_group='all'
        )

        # Добавляем изображение если есть
        if images_dir and article:
            self._add_product_image(product, article, images_dir)

        return product

    def _update_product(self, product, data, category, brand, animal_type):
        """Обновляет существующий товар"""
        product.name = data['name']
        product.category = category
        product.brand = brand
        product.price = self._parse_decimal(data.get('price', product.price))
        
        if data.get('weight'):
            product.weight = self._parse_decimal(data['weight'])
        
        if data.get('description'):
            product.description = data['description']

        # Обновляем статус наличия
        stock = data.get('stock')
        if stock is not None:
            try:
                stock_value = float(stock)
                product.stock_status = 'in_stock' if stock_value > 0 else 'out_of_stock'
            except (ValueError, TypeError):
                pass

        product.save()
        return product

    def _add_product_image(self, product, article, images_dir):
        """Добавляет изображение товара"""
        # Ищем файл изображения по артикулу
        extensions = ['.jpg', '.jpeg', '.png', '.webp']
        images_path = Path(images_dir)
        
        for ext in extensions:
            image_file = images_path / f"{article}{ext}"
            if image_file.exists():
                try:
                    with open(image_file, 'rb') as f:
                        product.image.save(
                            f"{article}{ext}",
                            File(f),
                            save=True
                        )
                    self.stdout.write(f'  → Добавлено изображение: {image_file.name}')
                    return
                except Exception as e:
                    self.stdout.write(self.style.WARNING(
                        f'  → Ошибка загрузки изображения {image_file}: {e}'
                    ))
                    return

    def _parse_decimal(self, value, default=0):
        """Парсит число в Decimal"""
        if value is None:
            return Decimal(default)
        try:
            # Заменяем запятую на точку (для белорусского формата)
            if isinstance(value, str):
                value = value.replace(',', '.')
            return Decimal(str(value))
        except (ValueError, TypeError):
            return Decimal(default)

