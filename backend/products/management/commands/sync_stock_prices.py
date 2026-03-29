"""
Команда для синхронизации остатков и цен из отчёта 1С «Оценка склада»
Формат: иерархический, вся текстовая информация в одной колонке (A),
уровни иерархии определяются отступами (пробелами в начале строки).

Логика расчёта цен:
  Штучные товары (шт):  Цена = Стоимость_в_рознице / Остаток
  Весовые товары (кг):  Цена_за_упаковку = (Стоимость_в_рознице / Остаток) × Вес_упаковки

Использование:
  python manage.py sync_stock_prices report.xlsx
  python manage.py sync_stock_prices report.xlsx --dry-run
  python manage.py sync_stock_prices report.xlsx --create-missing --update-brands
"""

from django.core.management.base import BaseCommand, CommandError
from django.utils.text import slugify
from products.models import Product, Brand, Category, AnimalType
import openpyxl
import re
from decimal import Decimal, ROUND_HALF_UP


class Command(BaseCommand):
    help = 'Синхронизация остатков и цен из отчёта «Оценка склада» (1С)'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Путь к файлу отчёта Excel')
        parser.add_argument(
            '--dry-run', action='store_true',
            help='Режим предпросмотра (без сохранения)',
        )
        parser.add_argument(
            '--create-missing', action='store_true',
            help='Создавать товары, которых нет в базе',
        )
        parser.add_argument(
            '--update-brands', action='store_true',
            help='Обновлять бренды товаров из иерархии отчёта',
        )

    def handle(self, *args, **options):
        file_path = options['file_path']
        dry_run = options['dry_run']
        create_missing = options['create_missing']
        update_brands = options['update_brands']

        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active
        except Exception as e:
            raise CommandError(f'Ошибка при чтении файла: {e}')

        if dry_run:
            self.stdout.write(self.style.WARNING('=== РЕЖИМ ПРЕДПРОСМОТРА ==='))

        col_cfg = self._detect_columns(sheet)
        self.stdout.write(f'Конфигурация колонок: {col_cfg}')

        default_category = None
        default_animal_type = None
        if create_missing:
            default_category, _ = Category.objects.get_or_create(
                slug='imported',
                defaults={'name': 'Импортированные товары',
                          'description': 'Из отчёта 1С'},
            )
            default_animal_type, _ = AnimalType.objects.get_or_create(
                slug='vse', defaults={'name': 'Все животные'},
            )

        stats = dict(updated=0, created=0, not_found=0, skipped=0, errors=0, total=0)
        current_brand_name = None

        for row_num in range(col_cfg['data_start_row'], sheet.max_row + 1):
            row = [cell.value for cell in sheet[row_num]]
            stats['total'] += 1

            try:
                rtype = self._detect_row_type(row, col_cfg)

                if rtype in ('warehouse', 'total', 'empty'):
                    stats['skipped'] += 1
                    continue

                if rtype == 'brand':
                    raw = str(row[col_cfg['text_col']]).strip() if row[col_cfg['text_col']] else None
                    if raw and not raw.startswith('<'):
                        current_brand_name = raw
                    else:
                        current_brand_name = None
                    stats['skipped'] += 1
                    continue

                if rtype == 'product':
                    self._process_product(
                        row, row_num, col_cfg,
                        current_brand_name,
                        dry_run, create_missing, update_brands,
                        stats, default_category, default_animal_type,
                    )
                else:
                    stats['skipped'] += 1

            except Exception as e:
                stats['errors'] += 1
                self.stdout.write(self.style.ERROR(f'Строка {row_num}: {e}'))

        self._print_summary(stats, dry_run)

    # ------------------------------------------------------------------
    #  Определение структуры колонок
    # ------------------------------------------------------------------
    def _detect_columns(self, sheet):
        cfg = {'text_col': 0}

        for row_num in range(1, min(20, sheet.max_row + 1)):
            for cell in sheet[row_num]:
                v = cell.value
                if not v or not isinstance(v, str):
                    continue
                vl = v.strip().lower()
                ci = cell.column - 1

                if 'остаток' in vl and 'склад' in vl:
                    cfg['stock_col'] = ci
                elif 'остаток' in vl:
                    cfg.setdefault('stock_col', ci)
                if 'себестоимость' in vl:
                    cfg['cost_col'] = ci
                if 'стоимость' in vl and 'розничн' in vl:
                    cfg['retail_col'] = ci

        cfg.setdefault('stock_col', 5)
        cfg.setdefault('cost_col', 6)
        cfg.setdefault('retail_col', 7)

        cfg['data_start_row'] = 11
        for row_num in range(8, min(25, sheet.max_row + 1)):
            v = sheet.cell(row=row_num, column=cfg['stock_col'] + 1).value
            if v is not None:
                try:
                    float(v)
                    cfg['data_start_row'] = row_num
                    break
                except (ValueError, TypeError):
                    pass

        return cfg

    # ------------------------------------------------------------------
    #  Определение типа строки по отступу и содержимому
    # ------------------------------------------------------------------
    def _detect_row_type(self, row, cfg):
        tc = cfg['text_col']
        raw = row[tc] if tc < len(row) and row[tc] is not None else None
        if raw is None:
            return 'empty'

        raw_str = str(raw)
        text = raw_str.strip()
        if not text:
            return 'empty'

        if 'итого' in text.lower():
            return 'total'

        indent = len(raw_str) - len(raw_str.lstrip())

        if indent > 0:
            return 'product'

        # indent == 0: отличаем склад, бренд и товар по содержимому
        low = text.lower()

        if any(k in low for k in ('магазин', 'чуп', 'ооо', 'ип ')):
            return 'warehouse'

        if text.startswith('<'):
            return 'brand'

        # Есть суффикс единицы измерения → товар
        if re.search(r',\s*,?\s*(шт|кг|г|л|мл)\s*$', text):
            return 'product'

        # Содержит числа с весовыми/объёмными единицами → товар
        if re.search(r'\d+[,.]?\d*\s*(кг|г|мл|л|таб)', text, re.IGNORECASE):
            return 'product'

        return 'brand'

    # ------------------------------------------------------------------
    #  Обработка строки товара
    # ------------------------------------------------------------------
    def _process_product(self, row, row_num, cfg, brand_name,
                         dry_run, create_missing, update_brands,
                         stats, default_category, default_animal_type):
        tc = cfg['text_col']
        raw = str(row[tc]).strip() if tc < len(row) and row[tc] else ''
        if len(raw) < 3:
            stats['skipped'] += 1
            return

        product_name, unit = self._parse_name_and_unit(raw)
        stock = self._num(row[cfg['stock_col']] if cfg['stock_col'] < len(row) else None)
        retail = self._num(row[cfg['retail_col']] if cfg['retail_col'] < len(row) else None)

        if stock is None or retail is None or stock <= 0 or retail <= 0:
            product = self._find_product(product_name)
            if product:
                if not dry_run:
                    product.stock_status = 'out_of_stock'
                    product.save(update_fields=['stock_status'])
                stats['updated'] += 1
            else:
                stats['not_found'] += 1
            return

        if unit == 'кг':
            pkg = self._extract_package_weight(raw)
            price_per_kg = Decimal(str(retail)) / Decimal(str(stock))
            new_price = (price_per_kg * Decimal(str(pkg))).quantize(
                Decimal('0.01'), rounding=ROUND_HALF_UP)
            weight_val = Decimal(str(pkg))
        else:
            new_price = (Decimal(str(retail)) / Decimal(str(stock))).quantize(
                Decimal('0.01'), rounding=ROUND_HALF_UP)
            weight_val = None

        product = self._find_product(product_name)

        if product:
            changes, fields = [], []

            if product.price != new_price:
                changes.append(f'цена: {product.price} -> {new_price} BYN')
                if not dry_run:
                    product.price = new_price
                    fields.append('price')

            if product.stock_status != 'in_stock':
                changes.append(f'статус: {product.stock_status} -> in_stock')
                if not dry_run:
                    product.stock_status = 'in_stock'
                    fields.append('stock_status')

            if product.unit != unit:
                changes.append(f'ед.изм: {product.unit} -> {unit}')
                if not dry_run:
                    product.unit = unit
                    fields.append('unit')

            if weight_val is not None and product.weight != weight_val:
                changes.append(f'вес: {product.weight} -> {weight_val}')
                if not dry_run:
                    product.weight = weight_val
                    fields.append('weight')

            if update_brands and brand_name:
                brand, _ = Brand.objects.get_or_create(
                    name=brand_name, defaults={'country': ''})
                if product.brand_id != brand.id:
                    changes.append(f'бренд -> {brand.name}')
                    if not dry_run:
                        product.brand = brand
                        fields.append('brand')

            if changes and not dry_run and fields:
                product.save(update_fields=fields)

            if changes:
                stats['updated'] += 1
                self.stdout.write(self.style.SUCCESS(
                    f'[OK] "{product_name[:55]}": {", ".join(changes)}'))
            else:
                stats['skipped'] += 1

        elif create_missing:
            brand = self._get_or_create_brand(brand_name)
            slug = slugify(product_name)[:200] or 'product'
            base_slug, ctr = slug, 1
            while Product.objects.filter(slug=slug).exists():
                slug = f'{base_slug}-{ctr}'
                ctr += 1

            if not dry_run:
                Product.objects.create(
                    name=product_name, slug=slug,
                    description='', composition='',
                    weight=weight_val if weight_val else Decimal('1'),
                    unit=unit, price=new_price,
                    category=default_category,
                    brand=brand,
                    animal_type=default_animal_type,
                    stock_status='in_stock', age_group='all',
                )
            stats['created'] += 1
            self.stdout.write(self.style.SUCCESS(
                f'[NEW] "{product_name[:55]}" | {new_price} BYN | {unit}'))
        else:
            stats['not_found'] += 1
            if stats['not_found'] <= 20:
                self.stdout.write(self.style.WARNING(
                    f'[!] Не найден: "{product_name[:65]}"'))

    # ------------------------------------------------------------------
    #  Вспомогательные методы
    # ------------------------------------------------------------------
    def _parse_name_and_unit(self, raw):
        name = ' '.join(raw.split())
        unit = 'шт'
        for u in ('шт', 'кг', 'г', 'л', 'мл'):
            pattern = rf',\s*,?\s*{u}\s*$'
            if re.search(pattern, name):
                unit = u
                name = re.sub(pattern, '', name).strip()
                break
        return name.rstrip(', '), unit

    def _extract_package_weight(self, raw):
        """Извлекает вес упаковки из названия для кг-товаров."""
        kg = re.search(r'(\d+[,.]?\d*)\s*кг', raw)
        if kg:
            w = float(kg.group(1).replace(',', '.'))
            if w > 0:
                return w

        g = re.search(r'(\d+[,.]?\d*)\s*г[.,\s]', raw)
        if g:
            w = float(g.group(1).replace(',', '.'))
            if w > 0:
                return w / 1000.0

        return 1.0

    def _find_product(self, name):
        product = Product.objects.filter(name__iexact=name).first()
        if product:
            return product

        words = name.split()[:4]
        if len(words) >= 2:
            product = Product.objects.filter(
                name__istartswith=' '.join(words)).first()
            if product:
                return product

        if len(name) > 10:
            product = Product.objects.filter(name__icontains=name[:30]).first()
            return product

        return None

    def _get_or_create_brand(self, brand_name):
        if not brand_name:
            brand_name = 'Неизвестный производитель'
        brand, _ = Brand.objects.get_or_create(
            name=brand_name, defaults={'country': ''})
        return brand

    @staticmethod
    def _num(val):
        if val is None:
            return None
        if isinstance(val, (int, float)):
            return float(val)
        if isinstance(val, str):
            try:
                return float(val.replace(' ', '').replace(',', '.'))
            except ValueError:
                return None
        return None

    def _print_summary(self, s, dry_run):
        self.stdout.write('\n' + '=' * 60)
        if dry_run:
            self.stdout.write(self.style.WARNING(
                'ПРЕДПРОСМОТР (изменения НЕ сохранены)'))
        else:
            self.stdout.write(self.style.SUCCESS('СИНХРОНИЗАЦИЯ ЗАВЕРШЕНА'))
        self.stdout.write(f'Всего строк: {s["total"]}')
        self.stdout.write(self.style.SUCCESS(f'Обновлено: {s["updated"]}'))
        if s['created']:
            self.stdout.write(self.style.SUCCESS(f'Создано: {s["created"]}'))
        if s['not_found']:
            self.stdout.write(self.style.WARNING(f'Не найдено: {s["not_found"]}'))
        if s['skipped']:
            self.stdout.write(f'Пропущено: {s["skipped"]}')
        if s['errors']:
            self.stdout.write(self.style.ERROR(f'Ошибок: {s["errors"]}'))
        self.stdout.write('=' * 60)
